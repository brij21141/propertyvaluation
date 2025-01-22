from django.shortcuts import render
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.models import User
import csv,json,time
from django.utils import timezone
from datetime import datetime,timedelta
from propval.models import UserDetails,Banks,Impdoc
from django.db.models import Count
from reception.models import ReceptionReport,ArchieveReceptionReport
from site_engineer.models import EngineerReport,ArchieveEngineerReport,HistoryEngineerReport,Floordetails,HistoryFloordetails,EngAttendance,EngDynamicdValue,HistoryEngDynamicdValue
from reporter.models import ReporterReport
from django.contrib.auth import login, authenticate, logout
from rest_framework.views import APIView
from rest_framework.response import Response
from .serializers import LoginSerializer,UserSerializer,EngineerSerializer,ReceptionSerializer,UserdetailSerializer,EngineerCreateSerializer,ReporterSerializer,UserProfileSerializer
from .serializers import ResetPasswordEmailRequestSerializer,ResetPasswordSerializer,BankSerializer,FileUploadSerializer,EngineerAttendanceSerializer
from rest_framework.authtoken.models import Token
from rest_framework.permissions import IsAuthenticated
from rest_framework import viewsets
from rest_framework.decorators import action 
from rest_framework.response import Response
from rest_framework import status
from rest_framework import generics
import openpyxl
from django.db.models import Q  
from django.dispatch import receiver
from django_rest_passwordreset.signals import reset_password_token_created
from django.core.mail import send_mail
from django.urls import reverse
from django_rest_passwordreset.models import ResetPasswordToken
import os
from django.conf import settings
from reception.models import Document
from django.db.models import Prefetch


# users=User.objects.all().filter(id__gt=1)
users=User.objects.all()
dictusers={
      "allusers":[]
}
for user in users:
        dictusers["allusers"].append([user.id,user.username,user.email,user.is_superuser,user.date_joined.strftime('%d-%m-%Y %a %H:%M:%S')]) 


def allUsers(request):
   
    return JsonResponse(dictusers)

def currentUser(request):
      if request.user.is_authenticated:
        useremail = request.user.email
        if(UserDetails.objects.filter(user_email=useremail).exists()):
            username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
            userrole = UserDetails.objects.get(user_email=useremail).role
            profileimg = UserDetails.objects.get(user_email=useremail).profileimage
            if(profileimg):
                  profileimgurl=profileimg.url
            else:
                  profileimgurl=None
            
            uid = UserDetails.objects.get(user_email=useremail).id
            return JsonResponse({'success': True, 'data':[uid,username,useremail,userrole,profileimgurl],'message': 'User details'})
      else:
        return JsonResponse({'success': False,'message': 'User not authenticated'})

class LoginAPI(APIView):
#      permission_classes = [IsAuthenticated]     ###use this line to only token authenticated user can use other apis ###
     def post(self, request):
          data = request.data
          print(data)
          serializer=LoginSerializer(data=data)
          if not serializer.is_valid():
                return Response({"status":False, "data":serializer.errors,"message":"Please provide a valid login"})
          username=serializer.data["username"]
          password=serializer.data["password"]
          user = authenticate(request, username=username,password=password)
          if user is not None:
            token , _ = Token.objects.get_or_create(user=user)
            login(request , user)
            # return JsonResponse({'success': True, 'data':str(token), 'message': 'User Authenticated successfully'})
            return JsonResponse({'success': True, 'message': 'User Authenticated successfully','token':str(token),'userid':user.id})
          else:
            return JsonResponse({'success': False,'message': 'User name or password is wrong'}) 
      #     return JsonResponse(data)
class LogoutAPI(APIView):
      def post(self,request):
            logout(request)
            return JsonResponse({'success': True,'message': 'User logged out successfully'})

def rolewiseuser(request):
      role_order = ['Admin', 'Reception', 'Engineer', 'Reporter']
      role_counts = UserDetails.objects.values('role').annotate(user_count=Count('id'))
      role_count_dict = {role: 0 for role in role_order}
      for entry in role_counts:
            role = entry['role']
            if role in role_order:
                  role_count_dict[role] = entry['user_count']
                  
             
      dictusersno={
            "usersno":[]
      }
      dictusersno["usersno"].append(role_count_dict['Admin'])
      dictusersno["usersno"].append(role_count_dict['Reception']) 
      dictusersno["usersno"].append(role_count_dict['Engineer']) 
      dictusersno["usersno"].append(role_count_dict['Reporter']) 

      return JsonResponse({"data":dictusersno['usersno']})

def export(request):
      response = HttpResponse(
            content_type ='text/csv',
            headers={'content-Disposition':'attachment; filename="Users.csv"'}
            
      )
      writer = csv.writer(response)
      queryset = User.objects.all().values(
            "id","username","email","is_superuser","date_joined"
      )
      for item in queryset:
            writer.writerow([str(item["id"]),item["username"],item["email"],item["is_superuser"],item["date_joined"]])
      return response
def bankexport(request):
      response = HttpResponse(
            content_type ='text/csv',
            headers={'content-Disposition':'attachment; filename="banks.csv"'}
            
      )
      writer = csv.writer(response)
      queryset = Banks.objects.all().values(
            "id","name","add1","add2","city","state","zip","contact_no","std","landline","email","website","gstin"
      )
      for item in queryset:
            writer.writerow([str(item["id"]),str(item["name"]),str(item["add1"]),str(item["add2"]),str(item["city"]),str(item["state"]),str(item["zip"]),str(item["contact_no"]),str(item["std"]),str(item["landline"]),str(item["email"]),str(item["website"]),str(item["gstin"])])
      return response
def receptiongenreportexport(request):
      # Create a workbook and add a worksheet  
    workbook = openpyxl.Workbook()  
    worksheet = workbook.active  

    # Add headers
#     worksheet.append(['Application Date', 'Application No.','Name','Phone Number' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter'])  # replace with actual headers  

    # Add data rows 
    if request.user.is_authenticated:
      useremail = request.user.email
      userid=request.user.id
      if(UserDetails.objects.filter(user_email=useremail).exists()):
            username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
            userrole = UserDetails.objects.get(user_email=useremail).role   
            if (userrole == "Admin" or userrole == "Reception"):
                  worksheet.append(['Application Date', 'Application No.','Name','Phone Number' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter']) 
                  for reception in ReceptionReport.objects.all(): 
                        worksheet.append([(reception.applicationdate).replace(tzinfo=None), reception.applicationnumber, reception.name, reception.phonenumber, reception.bankname, reception.bankvertical,
                           reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country, reception.visitingpersonname, reception.reportpersonname])  # replace with actual fields  
            else:
                  worksheet.append(['Application Date', 'Application No.','Name','Phone Number' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer']) 
                  for reception in ReceptionReport.objects.filter(visitingperson=userid).exclude(engineer='Submitted'): 
                        worksheet.append([(reception.applicationdate).replace(tzinfo=None), reception.applicationnumber, reception.name, reception.phonenumber, reception.bankname, reception.bankvertical,
                           reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country, reception.visitingpersonname, ])  # replace with actual fields  
    # Create response  
    response = HttpResponse(  
      #   content=workbook.stream,  
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  
        headers={'Content-Disposition': 'attachment; filename="ReceptionGeneratedReport.xlsx"'},  
    )  
    workbook.save(response)  
    return response

def engcompletedreportexport(request):
      # Create a workbook and add a worksheet  
    workbook = openpyxl.Workbook()  
    worksheet = workbook.active  

    # Add headers
#     worksheet.append(['Application Date', 'Application No.','Name','Phone Number' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter'])  # replace with actual headers  

    # Add data rows 
    if request.user.is_authenticated:
      useremail = request.user.email
      userid=request.user.id
      if(UserDetails.objects.filter(user_email=useremail).exists()):
            username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
            userrole = UserDetails.objects.get(user_email=useremail).role   
            if (userrole == "Admin" or userrole == "Reception"):
                  # worksheet.append([ 'Application No.','Name' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer']) 
                  worksheet.append(['Application No.','Name' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter', 'East Boundary', 'West Boundary', 'North Boundary', 'South Boundary', 'Ground Floor Area', 'First floor Area', 'Second Floor Area', 'Third Floor Area', 'Age of the property', 'Land Rate', 'Occupant', 'Rented', 'Land Mark', 'Road Width', 
                                    'High Tension line','Railway line','Nala','River','Pahad','Road comes under binding','Access issue','Other Checks','Other', 'Remarks','Priority']) 
                  for reception in EngineerReport.objects.all(): 
                        # worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                        #    reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname
                        #  ]) 
                        worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                        reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname,reception.receptionid.reportpersonname, reception.east, reception.west, reception.north, reception.south, reception.gfarea, reception.ffarea, reception.sfarea, reception.tfarea, reception.propertyage, reception.landrate, reception.Occupant, reception.rented, reception.landmark, reception.roadwidth
                        , reception.hightensionline, reception.railwayline, reception.nala, reception.river,reception.pahad, reception.roadcomesunderroadbinding, reception.propertyaccessissue,reception.othercheck, reception.others, reception.remark,reception.priority])
                        
                       
            else:
                  worksheet.append(['Application No.','Name' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter', 'East Boundary', 'West Boundary', 'North Boundary', 'South Boundary', 'Ground Floor Area', 'First floor Area', 'Second Floor Area', 'Third Floor Area', 'Age of the property', 'Land Rate', 'Occupant', 'Rented', 'Land Mark', 'Road Width', 
                                    'High Tension line','Railway line','Nala','River','Pahad','Road comes under binding','Access issue','Other Checks','Other', 'Remarks','Priority']) 
                  for reception in EngineerReport.objects.filter(Q(receptionid__visitingperson=userid) | (Q(receptionid__reportperson=userid  ) & ~Q(receptionid__reporter ='Submitted'))): 
                        # worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                        #    reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname
                        #  ]) 
                        worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                        reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname,reception.receptionid.reportpersonname, reception.east, reception.west, reception.north, reception.south, reception.gfarea, reception.ffarea, reception.sfarea, reception.tfarea, reception.propertyage, reception.landrate, reception.Occupant, reception.rented, reception.landmark, reception.roadwidth
                        , reception.hightensionline, reception.railwayline, reception.nala, reception.river,reception.pahad, reception.roadcomesunderroadbinding, reception.propertyaccessissue,reception.othercheck, reception.others, reception.remark,reception.priority])
                  
    # Create response  
    response = HttpResponse(  
      #   content=workbook.stream,  
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  
        headers={'Content-Disposition': 'attachment; filename="EngineerCompletedReport.xlsx"'},  
    )  
    workbook.save(response)  
    return response
def reportercompletedexport(request):
      # Create a workbook and add a worksheet  
    workbook = openpyxl.Workbook()  
    worksheet = workbook.active  

    # Add headers
#     worksheet.append(['Application Date', 'Application No.','Name','Phone Number' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter'])  # replace with actual headers  

    # Add data rows 
    if request.user.is_authenticated:
      useremail = request.user.email
      userid=request.user.id
      if(UserDetails.objects.filter(user_email=useremail).exists()):
            username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
            userrole = UserDetails.objects.get(user_email=useremail).role   
            if (userrole == "Admin" or userrole == "Reception"):
                  # worksheet.append([ 'Application No.','Name' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer']) 
                  worksheet.append(['Application No.','Name' ,'Bank Name', 'Bank Vertical', 'Address line 1', 'Address Line 2','City', 'State','Pin code','Country','Assigned Engineer','Assigned Reporter', 'East Boundary', 'West Boundary', 'North Boundary', 'South Boundary', 'Ground Floor Area', 'First floor Area', 'Second Floor Area', 'Third Floor Area', 'Age of the property', 'Land Rate', 'Occupant', 'Rented', 'Land Mark', 'Road Width', 
                                    'High Tension line','Railway line','Nala','River','Pahad','Road comes under binding','Access issue','Other Checks','Other', 'Remarks','Priority']) 
                  for reception in EngineerReport.objects.all(): 
                        worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                        reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname,reception.receptionid.reportpersonname, reception.east, reception.west, reception.north, reception.south, reception.gfarea, reception.ffarea, reception.sfarea, reception.tfarea, reception.propertyage, reception.landrate, reception.Occupant, reception.rented, reception.landmark, reception.roadwidth
                        , reception.hightensionline, reception.railwayline, reception.nala, reception.river,reception.pahad, reception.roadcomesunderroadbinding, reception.propertyaccessissue,reception.othercheck, reception.others, reception.remark,reception.priority])
                        
                       
            else:
                  # worksheet.append(['Inspection Date','Application No.','Name','Document Holder Name' ,'Property Type', 'Address', 'Address line 2','City','State','Pin code','Country','Landmark'
                  #                   ,'Legal Address', 'Legal Address line 2','City','State','Pin code','Country','Ward Land no.','Approach Road width','Vicinity','Property Location','Property Identification'
                  #                   ,'Connectivity Infra','Nearest railway station','Nearest bus stop','Nearest hospital','Nearest landmark', 'Property Usage type ', 'Additional Amminities', 'Legal status of property', 'Premises type', 'Taxation Maintainence cost'
                  #                   ,'Renting Potential','Market Rental','Occupied by','property rented','Occupant list','As per Doc.East boundry'
                  #                   ,'As per Doc.West boundry','As per Doc.North boundry','As per Doc.South boundry','As per Site.East boundry'
                  #                   ,'As per Site.West boundry','As per Site.North boundry','As per Site.South boundry','Structure Type','No. of Floor','No. of Wings','No.of Unit Each Flooe'
                  #                   ,'No.of Lift Each Wings',reception.ageproperty,reception.futurelife,reception.exterior,reception.internalcomposition,reception.constructionquality
                  #                   ,reception.beamcolumnstru,reception.commonarearemark,reception.otherobservation,reception.floornfinish,reception.roofingnterracing,reception.nooflifts
                  #                   ,reception.qualityfixing,reception.constasperaprove,reception.aprnodate,reception.constnodate,reception.violationifany,reception.confirmlocalbilaws
                  #                   ,reception.otherverifieddoc,reception.carpetareaflwise,reception.aprovbuaflwise,reception.govtapprovalrate,reception.recomendrate,reception.mktvalueinfig
                  #                   ,reception.mktvalueinwords,reception.forcessalesless,reception.arearatepsft,reception.areavalueinfig,reception.areavalueinwords,reception.landarea
                  #                   ,reception.govtaprovelandrate,reception.recommendedlandrate,reception.landvalue,reception.actualbua,reception.asperpermissionbua,reception.constcostperaminity
                  #                   ,reception.totalconstvalue,reception.depreciatedconstvalue,reception.landvaluedepndepconstvalueinfig,reception.landvaluedepndepconstvalueinword,reception.forcesalevaluepersqft
                  #                   ,reception.forcesvalueinfig,reception.forcesvalueinword,reception.marketibility,reception.valuationresult,reception.remark,reception.lat,reception.lng
                  #                   ,reception.placeid,(reception.datecreated.replace(tzinfo=None)),(reception.updated_at).replace(tzinfo=None) ]) 
                  for reception in ReporterReport.objects.filter(receptionid__reportperson=userid): 
                  #       worksheet.append([ reception.applicationnumber, reception.name, reception.bankname, reception.casetype,
                  #       reception.add1, reception.add2, reception.city, reception.region, reception.zip, reception.country,reception.receptionid.visitingpersonname,reception.receptionid.reportpersonname, reception.east, reception.west, reception.north, reception.south, reception.gfarea, reception.ffarea, reception.sfarea, reception.tfarea, reception.propertyage, reception.landrate, reception.Occupant, reception.rented, reception.landmark, reception.roadwidth
                  #       , reception.hightensionline, reception.railwayline, reception.nala, reception.river,reception.pahad, reception.roadcomesunderroadbinding, reception.propertyaccessissue,reception.othercheck, reception.others, reception.remark,reception.priority])
                        row_data = [  
                                    (reception.inspectiondate).replace(tzinfo=None),reception.applicationnumber,reception.name,reception.docholdername
                                    ,reception.propertytype,reception.add1,reception.add2,reception.city,reception.region
                                    ,reception.zip,reception.country,reception.landmark,reception.ladd1,reception.ladd2
                                    ,reception.lcity,reception.lregion,reception.lzip,reception.lcountry,reception.wardlandno
                                    ,reception.approachroadwidth,reception.vicinity,reception.propertylocation,reception.propertyidentification
                                    ,reception.connectivityinfrastructure,reception.railwaystation,reception.busstop,reception.hospital,reception.nearestlandmark
                                    ,reception.typeusageproperty,reception.additionalamenities,reception.legalstatusproperty,reception.typepremises,reception.taxationmaintancecost
                                    ,reception.rentingpotential,reception.marketrentals,reception.occupiedby,reception.ispropertyrented,reception.listofoccupants,reception.perdcboundryeast
                                    ,reception.perdcboundrywset,reception.perdcboundrynorth,reception.perdcboundrysouth,reception.perstboundryeast,reception.perstboundrywest,reception.perstboundrynorth
                                    ,reception.perstboundrysouth,reception.typestructure,reception.nofloors,reception.nowings,reception.nouniteachfloor
                                    ,reception.nolifteachwing,reception.ageproperty,reception.futurelife,reception.exterior,reception.internalcomposition,reception.constructionquality
                                    ,reception.beamcolumnstru,reception.commonarearemark,reception.otherobservation,reception.floornfinish,reception.roofingnterracing,reception.nooflifts
                                    ,reception.qualityfixing,reception.constasperaprove,reception.aprnodate,reception.constnodate,reception.violationifany,reception.confirmlocalbilaws
                                    ,reception.otherverifieddoc,reception.carpetareaflwise,reception.aprovbuaflwise,reception.govtapprovalrate,reception.recomendrate,reception.mktvalueinfig
                                    ,reception.mktvalueinwords,reception.forcessalesless,reception.arearatepsft,reception.areavalueinfig,reception.areavalueinwords,reception.landarea
                                    ,reception.govtaprovelandrate,reception.recommendedlandrate,reception.landvalue,reception.actualbua,reception.asperpermissionbua,reception.constcostperaminity
                                    ,reception.totalconstvalue,reception.depreciatedconstvalue,reception.landvaluedepndepconstvalueinfig,reception.landvaluedepndepconstvalueinword,reception.forcesalevaluepersqft
                                    ,reception.forcesvalueinfig,reception.forcesvalueinword,reception.marketibility,reception.valuationresult,reception.remark,reception.lat,reception.lng
                                    ,reception.placeid,(reception.datecreated.replace(tzinfo=None)),(reception.updated_at).replace(tzinfo=None) 
                              ]  
                        worksheet.append(row_data)    
                  
    # Create response  
    response = HttpResponse(  
      #   content=workbook.stream,  
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',  
        headers={'Content-Disposition': 'attachment; filename="EngineerCompletedReport.xlsx"'},  
    )  
    workbook.save(response)  
    return response

def recepreportpriority(request,uid):
    try:
        u=ReceptionReport.objects.get(pk=uid)
        if u.priority:
            u.priority=False
            msg='Report has been changed to normal report successfully'
        else:
            u.priority=True
            msg='Report Priority has been changed successfully'
        u.save()
        return JsonResponse({'success': True,'message': msg})
    except ReceptionReport.DoesNotExist:
        # return redirect('home')
        return JsonResponse({'success': False, 'error':'Priority not set yet'})
def recepreportnpa(request,uid):
    try:
        u=ReceptionReport.objects.get(pk=uid)
        if u.npa:
            u.npa=False
            msg='Report has been changed to normal report successfully'
            npacase=False
            
        else:
            u.npa=True
            msg='Report has been changed to npa report successfully'
            npacase=True
        u.save()
        return JsonResponse({'success': True,'message': msg,'npacase': npacase})
    except ReceptionReport.DoesNotExist:
        # return redirect('home')
        return JsonResponse({'success': False, 'error':'npa not set yet'})

def engreportstatus(request,status,uid):
      # print (status)
      try:
            msg = 'for reception dashboard'
            messg='engineer report summary'
            # print (uid)
            if(request.method == 'PUT'):
                  if(uid>0 and status=='S'):
                        data = json.loads(request.body)
                        try:
                              er=ReceptionReport.objects.get(pk=uid)
                              if er.engineer is None:
                                    print ("hhhhh")
                                    er.engineer='InProgress'
                                    er.engineerholdcause=''
                                    msg='InProgress'
                              elif er.engineer=='InProgress':
                                    er.engineer='Hold'
                                    er.engineerholdcause=data['cause']
                                    msg='Hold'
                              else:
                                    print ("else---")
                                    er.engineer=None
                                    er.engineerholdcause=''
                                    msg='Pending'
                              messg = 'Report Status has been changed successfully'
                              er.save()
                        except ReceptionReport.DoesNotExist:
                              return JsonResponse({'error': 'Item not found'}, status=404)  
            else:
                       if(uid>0 and status=='S'):
                        try:
                              er=ReceptionReport.objects.get(pk=uid)
                              if er.engineer is None:
                                    print ("hhhhh")
                                    er.engineer='InProgress'
                                    er.engineerholdcause=''
                                    msg='InProgress'
                              elif er.engineer=='InProgress':
                                    er.engineer='Hold'
                                    er.engineerholdcause=data['cause']
                                    msg='Hold'
                              else:
                                    print ("else---")
                                    er.engineer=None
                                    er.engineerholdcause=''
                                    msg='Pending'
                              messg = 'Report Status has been changed successfully'
                              er.save()
                        except ReceptionReport.DoesNotExist:
                              return JsonResponse({'error': 'Item not found'}, status=404)  
            if request.user.is_authenticated:
                  useremail = request.user.email
                  if(UserDetails.objects.filter(user_email=useremail).exists()):
                        username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
                        userrole = UserDetails.objects.get(user_email=useremail).role 
                        if(status =='S'):
                              uid = request.user.id
                        
                        if (userrole == "Admin" or userrole == "Reception"):
                              # receivedrequest=ReceptionReport.objects.exclude(engineer ='Submitted')
                              totalrequestnumber = ReceptionReport.objects.filter(Q(reportperson__gt=0)|Q(visitingperson__gt=0)).count()
                              totalcompleted = ReceptionReport.objects.filter(engineer='Submitted').count()
                              inprogress = ReceptionReport.objects.filter(engineer='InProgress').count()
                              hold = ReceptionReport.objects.filter(engineer='Hold').count()
                              # pendingrequest = totalrequestnumber-(totalcompleted+inprogress+hold)
                              pendingrequest = totalrequestnumber-(totalcompleted+hold)
                              
                        else:
                              # receivedrequest=ReceptionReport.objects.exclude(engineer ='Submitted')
                              # totalrequestnumber = ReceptionReport.objects.filter(visitingpersonname=username).count()
                              totalrequestnumber = ReceptionReport.objects.filter(visitingperson=uid).count()
                              print(uid, totalrequestnumber)
                              totalcompleted = ReceptionReport.objects.filter(visitingperson=uid, engineer='Submitted').count()
                              inprogress = ReceptionReport.objects.filter(visitingperson=uid, engineer='InProgress').count()
                              hold = ReceptionReport.objects.filter(visitingperson=uid, engineer='Hold').count()
                              # pendingrequest = totalrequestnumber-(totalcompleted+inprogress+hold)
                              pendingrequest = totalrequestnumber-(totalcompleted+hold)
                              
            dictengstatus={
                  "engstatus":[]
            }
            dictengstatus["engstatus"].append(totalrequestnumber)
             
            dictengstatus["engstatus"].append(pendingrequest) 
            dictengstatus["engstatus"].append(inprogress) 
            dictengstatus["engstatus"].append(hold)
            dictengstatus["engstatus"].append(totalcompleted)            

            return JsonResponse({'success': True,'message': messg,
                                 'msg':msg,
                                 'tot':totalrequestnumber,
                                 'com':totalcompleted,
                                 'pend':pendingrequest,
                                 'inprog':inprogress,
                                 'hold':hold,
                                 "data":dictengstatus['engstatus']})
      except EngineerReport.DoesNotExist:
            return JsonResponse({'success': False, 'error':'Status not changed yet'})

def reporterreportstatus(request,uid):
    # print (repid)
      msg = 'for reception dashboard'
      messg='reporter report summary'
      try:
            if uid>0:
                  er=ReceptionReport.objects.get(pk=uid)
                  if er.reporter is None:
                        er.reporter='InProgress'
                        msg='InProgress'
                  elif er.reporter=='InProgress':
                        er.reporter='Hold'
                        msg='Hold'
                  else:
                        er.reporter=None
                        msg='Pending'
                  messg='Report Status has been changed successfully'
                  er.save()

            if request.user.is_authenticated:
                  useremail = request.user.email
                  if(UserDetails.objects.filter(user_email=useremail).exists()):
                        username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
                        userrole = UserDetails.objects.get(user_email=useremail).role   
            # print(username)
            # results = ReceptionReport.objects.filter(visitingpersonname=username).values('engineer').annotate(count=Count('id'))
            # print(results)
                        if (userrole == "Admin" or userrole == "Reception"):
                              # receivedrequest=ReceptionReport.objects.exclude(engineer ='Submitted')
                              totalrequestnumber = ReceptionReport.objects.filter(Q(reportperson__gt=0)|Q(visitingperson__gt=0)).count()
                              # totalrequestnumber = ReceptionReport.objects.count()
                              totalcompleted = ReceptionReport.objects.filter(reporter='Submitted').count()
                              inprogress = ReceptionReport.objects.filter(reporter='InProgress').count()
                              hold = ReceptionReport.objects.filter(reporter='Hold').count()
                              # pendingrequest = totalrequestnumber-(totalcompleted+inprogress+hold)
                              pendingrequest = totalrequestnumber-(totalcompleted+hold)
                        else:
                              # receivedrequest=ReceptionReport.objects.exclude(engineer ='Submitted')
                              # totalrequestnumber = ReceptionReport.objects.filter(visitingpersonname=username).count()
                              totalrequestnumber = ReceptionReport.objects.filter(reportpersonname=username).count()
                              totalcompleted = ReceptionReport.objects.filter(reportpersonname=username, reporter='Submitted').count()
                              inprogress = ReceptionReport.objects.filter(reportpersonname=username, reporter='InProgress').count()
                              hold = ReceptionReport.objects.filter(reportpersonname=username, reporter='Hold').count()
                              print(hold)
                              # pendingrequest = totalrequestnumber-(totalcompleted+inprogress+hold)
                              pendingrequest = totalrequestnumber-(totalcompleted+hold)
                              # print(totalrequestnumber-totalcompleted)
            
            dictrepstatus={
                  "repstatus":[]
            }
            dictrepstatus["repstatus"].append(totalrequestnumber)
            
            dictrepstatus["repstatus"].append(pendingrequest) 
            dictrepstatus["repstatus"].append(inprogress) 
            dictrepstatus["repstatus"].append(hold) 
            dictrepstatus["repstatus"].append(totalcompleted) 

            return JsonResponse({'success': True,'message': messg,
                                 'msg':msg,
                                 'tot':totalrequestnumber,
                                 'com':totalcompleted,
                                 'pend':pendingrequest,
                                 'inprog':inprogress,
                                 'hold':hold,
                                 "data":dictrepstatus['repstatus']})
      except EngineerReport.DoesNotExist:
            return JsonResponse({'success': False, 'error':'Status not changed yet'})
def reportassign(request,uid):
    try:
      #   recid=EngineerReport.objects.get(pk=uid).receptionid_id
      #   u=ReceptionReport.objects.get(pk=recid)
        u=ReceptionReport.objects.get(pk=uid)
        msg='assigned'
        if request.user.is_authenticated:
            userid = request.user.id
            useremail = request.user.email
        if(UserDetails.objects.filter(user_email=useremail).exists()):
            username = UserDetails.objects.get(user_email=useremail).first_name+" "+UserDetails.objects.get(user_email=useremail).last_name
                  
            if u.reportperson == 0 or u.reportperson == '':
                  if UserDetails.objects.filter(user_id=userid,role="Reporter").exists():
                        u.reportperson=userid
                        u.reportpersonname=username
                        msg='assigned'
                  else: 
                        tot=ReceptionReport.objects.filter(reportperson=0).count()
                        msg = 'Reporter not found. You are not allowed to assign as you are not a reporter'
                        return JsonResponse({'success': False,'message': f'Report {msg} ','msg': msg,'tot': tot})
            elif u.reportperson != userid:
                 msg = 'Already assigned to someone else'
                 
            else:
                  u.reportperson=0
                  u.reportpersonname = 'Reporter Not Assigned'
                  msg='Unassigned'
            u.save()
            tot=ReceptionReport.objects.filter(reportperson=0).count()
        return JsonResponse({'success': True,'message': f'Report {msg} ','msg': msg,'tot': tot})
    except ReceptionReport.DoesNotExist:
        # return redirect('home')
        return JsonResponse({'success': False, 'error':'Report could not be assigned'})
    
#Apis
class EngAttendanceViewSet(viewsets.ModelViewSet):
     queryset = EngAttendance.objects.all()
     serializer_class=EngineerAttendanceSerializer
     @action(detail=False, methods=['put'])
     def updateout(self, request, *args, **kwargs):  
      #   id = kwargs.get('pk')  # Use 'pk' for the primary key in DRF  
        appno = request.data.get('applicationnumber')  
        recpid = request.data.get('receptionid')  
        instance = self.get_queryset().filter(applicationnumber=appno, receptionid_id=recpid).order_by('-indatetinme').first()  

        if instance is None:  
            return Response({'detail': 'Record not found'}, status=404)  

        instance.outdatetinme = timezone.now()  
        instance.save(update_fields=['outdatetinme'])  # Only update the outdatetime field  
        serializer = self.get_serializer(instance)  
        return Response(serializer.data)  

class EngineerViewSet(viewsets.ModelViewSet):
     queryset = EngineerReport.objects.all()
#      for report in queryset:
#             print(report)
     serializer_class=EngineerSerializer

#http://127.0.0.1:8000/api/engineer/createengreport/     adding record of eng report api
     @action(detail=False, methods=['post'])
     def createengreport(self, request, pk=None):
      #     queryset = EngineerReport.objects.all()
          serializer_class=EngineerCreateSerializer(data=request.data,context={'request':request})
          if serializer_class.is_valid():  
            serializer_class.save()  
            print(request.data.get('receptionid').split('/')[-2] ) 
            queryset=ReceptionReport.objects.get(id=request.data.get('receptionid').split('/')[-2])
            print(queryset)
            queryset.engineer='Submitted'
            queryset.save()
            # rr=ReceptionReport.objects.get(applicationnumber=app_number,pk=repid)
        
            return Response({'success': True, 'data': serializer_class.data}, status=status.HTTP_201_CREATED)  
          return Response(serializer_class.errors, status=status.HTTP_400_BAD_REQUEST)  
 
#http://127.0.0.1:8000/api/engineer/updateengreport/     updating record of eng report api
     @action(detail=True, methods=['put'])
     def updateengreport(self, request, pk=None):
          queryset = EngineerReport.objects.get(pk=pk)
          serializer_class=EngineerCreateSerializer(queryset,data=request.data,context={'request':request})
          if serializer_class.is_valid():  
            serializer_class.save()  
            return Response({'success': True, 'message':f'Record id: {pk} Application#: {queryset.applicationnumber} updated successfully','data': serializer_class.data}, status=status.HTTP_202_ACCEPTED)  
          return Response(serializer_class.errors, status=status.HTTP_400_BAD_REQUEST)  
          
#http://127.0.0.1:8000/api/engineer/engineercomjob/
     @action(detail=False,methods=['get'])
     def engineercomjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid')  
          engineercompjob_serializer=EngineerSerializer(engineercompjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineercompjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engcomjobpdfarctoo/   api to create engineer pdf also for archive data
     @action(detail=True,methods=['get'])
     def engcomjobpdfarctoo(self, request,pk=None):
          try:
            working_records = EngineerReport.objects.filter(pk=pk) 
            if working_records.exists():
                 engcomjobpdfarctoo_serializer=EngineerSerializer(working_records,many=True,context={'request':request})
            else:
                 archive_records = ArchieveEngineerReport.objects.filter(pk=pk) 
                 if archive_records.exists():
                  engcomjobpdfarctoo_serializer=EngineerSerializer(archive_records,many=True,context={'request':request})
             
          except EngineerReport.DoesNotExist:
            return JsonResponse({'success': False, 'error':'Record not found'})
      #     engcomjobpdfarctoo=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid')  
      #     engcomjobpdfarctoo_serializer=EngineerSerializer(engcomjobpdfarctoo,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engcomjobpdfarctoo_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engineercomjobapi/   for mobile app
     @action(detail=True,methods=['get'])
     def engineercomjobapi(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted',userdetailsid_id=pk )  
          engineercompjob_serializer=EngineerCreateSerializer(engineercompjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineercompjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engineerpendjob/     
     @action(detail=False,methods=['get'])
     def engineerpendjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          engineerpendjob=ReceptionReport.objects.filter(Q(engineer=None) | Q(engineer='null') | Q(engineer='InProgress'))
          engineerpendjob_serializer=ReceptionSerializer(engineerpendjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineerpendjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engineerinprogjob/     
     @action(detail=False,methods=['get'])
     def engineerinprogjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          engineerinprogjob=ReceptionReport.objects.filter(engineer='InProgress')  
          engineerinprogjob_serializer=ReceptionSerializer(engineerinprogjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineerinprogjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engineerholdjob/     
     @action(detail=False,methods=['get'])
     def engineerholdjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          engineerholdjob=ReceptionReport.objects.filter(engineer='Hold')  
          engineerholdjob_serializer=ReceptionSerializer(engineerholdjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineerholdjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/engineerrecjob/     
     @action(detail=False,methods=['get'])
     def engineerrecjob(self, request,pk=None):
          engineerrecjob=ReceptionReport.objects.all()  
          engineerrecjob_serializer=ReceptionSerializer(engineerrecjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': engineerrecjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/engineer/25/engreportstatus       
     @action(detail=True,methods=['get'])
     def engreportstatus(self,request,pk):
      
      try:
            messg='engineer report summary'
            totalrequestnumber = ReceptionReport.objects.filter(visitingperson=pk).count()
            # print(uid, totalrequestnumber)
            totalcompleted = ReceptionReport.objects.filter(visitingperson=pk, engineer='Submitted').count()
            inprogress = ReceptionReport.objects.filter(visitingperson=pk, engineer='InProgress').count()
            hold = ReceptionReport.objects.filter(visitingperson=pk, engineer='Hold').count()
            # pendingrequest = totalrequestnumber-(totalcompleted+inprogress+hold)
            pendingrequest = totalrequestnumber-(totalcompleted+hold)
            return JsonResponse({'success': True,'message': messg,
                                 'tot':totalrequestnumber,
                                 'com':totalcompleted,
                                 'pend':pendingrequest,
                                 'inprog':inprogress,
                                 'hold':hold})
      except EngineerReport.DoesNotExist:
            return JsonResponse({'success': False, 'error':'Status not changed yet'})
     
class UserProfileUpdateView(generics.UpdateAPIView):
     queryset = UserDetails.objects.all()
     serializer_class=UserProfileSerializer
#      permission_classes = [IsAuthenticated] 

#      def get_object(self):  
#         return self.request.user

class BankViewSet(viewsets.ModelViewSet):
     queryset = Banks.objects.all()
     serializer_class=BankSerializer

class ReporterViewSet(viewsets.ModelViewSet):
     queryset = ReporterReport.objects.all()
     serializer_class=ReporterSerializer

# http://127.0.0.1:8000/api/reporter/reporterrecjob/     addin record of eng report api
     @action(detail=False, methods=['get'])
     def reporterrecjob(self, request, pk=None):
          queryset = ReceptionReport.objects.all()
          queryset_serializer=ReceptionSerializer(queryset,many=True,context={'request':request})
          return Response({'success': True, 'data': queryset_serializer.data}, status=status.HTTP_200_OK)  
          
#http://127.0.0.1:8000/api/reporter/reportercomjob/
     @action(detail=False,methods=['get'])
     def reportercomjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          reportercomjob=ReporterReport.objects.filter(receptionid__reporter='Submitted').select_related('receptionid').select_related('bankid')  
          reportercomjob_serializer=ReporterSerializer(reportercomjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': reportercomjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/reporter/reporterpendjob/     
     @action(detail=False,methods=['get'])
     def reporterpendjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          reporterpendjob=ReceptionReport.objects.filter(Q(reporter=None) | Q(reporter='null') | Q(reporter='InProgress'))
          reporterpendjob_serializer=ReceptionSerializer(reporterpendjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': reporterpendjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/reporter/reporterinprogjob/     
     @action(detail=False,methods=['get'])
     def reporterinprogjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          reporterinprogjob=ReceptionReport.objects.filter(reporter='InProgress')  
          reporterinprogjob_serializer=EngineerSerializer(reporterinprogjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': reporterinprogjob_serializer.data
             }, status=status.HTTP_200_OK)
#http://127.0.0.1:8000/api/reporter/reporterholdjob/     
     @action(detail=False,methods=['get'])
     def reporterholdjob(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          reporterholdjob=ReceptionReport.objects.filter(reporter='Hold')  
          reporterholdjob_serializer=ReceptionSerializer(reporterholdjob,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': reporterholdjob_serializer.data
             }, status=status.HTTP_200_OK)

class UserViewSet(viewsets.ModelViewSet):
     queryset = User.objects.all()
     serializer_class=UserSerializer

class UserdetailViewSet(viewsets.ModelViewSet):
     queryset = UserDetails.objects.all()
     serializer_class=UserdetailSerializer

     @action(detail=True,methods=['get'])
     def getprofiledetail(self, request,pk=None):
      #     engineercompjob=EngineerReport.objects.filter(receptionid__engineer='Submitted').select_related('receptionid').values('receptionid__visitingpersonname', 'receptionid__phonenumber')
          userprofile=UserDetails.objects.filter(user_id=pk) 
          userprofile_serializer=UserdetailSerializer(userprofile,many=True,context={'request':request})
      #     return Response(engineercompjob_serializer.data)
          return Response({
            'success': True,
            'data': userprofile_serializer.data
             }, status=status.HTTP_200_OK)

class ReceptionViewSet(viewsets.ModelViewSet):
     permission_classes = [IsAuthenticated]
     queryset = ReceptionReport.objects.all()
     serializer_class=ReceptionSerializer

     @action(detail=True,methods=['get'])
     def engineerjob(self, request,pk=None):
          engineerjob=ReceptionReport.objects.filter(visitingperson=pk).exclude(engineer ='Submitted')
          reception_serializer=ReceptionSerializer(engineerjob,many=True,context={'request':request})
          print("getting reception report for a particular engineer" , pk, "engineer")
          return Response(reception_serializer.data)
     @action(detail=False,methods=['get'])
     def reporterunassigned(self, request,pk=None):
          reporterunassigned=ReceptionReport.objects.filter(reportperson=0)
          reception_serializer=ReceptionSerializer(reporterunassigned,many=True,context={'request':request})
          return Response({
               'success':True,
               'data':reception_serializer.data,
               },status=status.HTTP_200_OK)
     
class ResetPasswordRequestView(generics.GenericAPIView):
    serializer_class = ResetPasswordEmailRequestSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        email = serializer.validated_data['email']
        user = User.objects.filter(email=email).first()
        if user:
            reset_password_token_created.send(sender=self.__class__, instance=user, reset_password_token=ResetPasswordToken.objects.create(user=user))
            
        return Response({"message": "Password reset link has been sent to your email."}, status=status.HTTP_200_OK)

class ResetPasswordConfirmView(generics.GenericAPIView):
    serializer_class = ResetPasswordSerializer

    def post(self, request):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        token = serializer.validated_data['token']
        password = serializer.validated_data['password']
        reset_password_token = ResetPasswordToken.objects.filter(key=token).first()
        if reset_password_token and not reset_password_token.is_expired():
            user = reset_password_token.user
            user.set_password(password)
            user.save()
            reset_password_token.delete()
            return Response({"message": "Password has been reset successfully."}, status=status.HTTP_200_OK)
        return Response({"message": "Invalid or expired token."}, status=status.HTTP_400_BAD_REQUEST)
class DocumentUploadView(APIView):  
    def post(self, request,recid):  
        # Extract common metadata from the request  
      #   serializer = DocumentSerializer(data=request.data)  
      #   if serializer.is_valid(): 
        application_number = request.data.get('application_number')  
        reception_idno = request.data.get('reception_idno')  
        username = request.data.get('username')  
        role = request.data.get('role')  
        usersdetailsid = request.data.get('usersdetailsid')  
        platform = request.data.get('platform')  

        # Check if file data is provided  
        if not request.FILES:  
            return Response({'message': 'No files provided.'}, status=status.HTTP_400_BAD_REQUEST)  

        # Handle the file uploads  
        for file_data in request.FILES.getlist('files'):  
            upload_dir = os.path.join(settings.MEDIA_ROOT, 'engineer')  
            os.makedirs(upload_dir, exist_ok=True)  

            # Define the full path for the file  
            newfilename = application_number + "_"+str(recid)+"_"+str(time.time())+'_' + file_data.name
            file_path = os.path.join(upload_dir, newfilename)  
            # Save the uploaded file  
            with open(file_path, 'wb+') as destination:  
                for chunk in file_data.chunks():  
                    destination.write(chunk)  

            # Create a new Document instance for each file  
            Document.objects.create(  
                application_number=application_number,  
                reception_idno=reception_idno,  
                username=username,  
                role=role,  
                usersdetailsid=usersdetailsid,  
                platform=platform,  
            #     file_path=os.path.join('engineer', file_data.name),  # Save the relative path  
            #     file_path=file_path,  
                file_path='engineer/'+newfilename,  
                file_name=newfilename,  
            )  

        return Response({'message': 'Files uploaded successfully.'}, status=status.HTTP_201_CREATED) 
    
class DocumentgetView(APIView):
     def get(self, request, appno,recid):
          try:
            queryset =  Document.objects.filter(application_number=appno,reception_idno=recid, platform = 'engineer')
            if queryset.exists():
                  serializer = FileUploadSerializer(queryset, many=True,context={'request':request})
                  return JsonResponse({'success':True, 'message':"file/files found",'data':serializer.data}, status=status.HTTP_200_OK)
            else:
                 return JsonResponse({'success':True, 'message':"no file found"}, status=status.HTTP_200_OK)
          except Exception as e:
            return JsonResponse({'success':False,'error': str(e)}, status=status.HTTP_400_BAD_REQUEST)  
            

def homedatefilter(request):
      stdate_str = request.GET.get('startdate')
      endate_str = request.GET.get('enddate')
      bankid = request.GET.get('bankid')
      engstatus = request.GET.get('engstatus')
      repstatus = request.GET.get('repstatus')
      vertical = request.GET.get('vertical')
      
      # Add one day to endate
      
      # endate = datetime.combine(endate, time.max)
      if stdate_str:
           stdate = datetime.strptime(stdate_str, "%Y-%m-%d").date()
           searchresults=ReceptionReport.objects.filter(datecreated__gte=stdate)
           print("A")
      if endate_str:
           endate = datetime.strptime(endate_str, "%Y-%m-%d").date()
           endate = endate + timedelta(days=1)
           searchresults=ReceptionReport.objects.filter(datecreated__lte=endate)
           print("B")
      if (stdate_str and endate_str):
           stdate = datetime.strptime(stdate_str, "%Y-%m-%d").date()
           endate = datetime.strptime(endate_str, "%Y-%m-%d").date()
           endate = endate + timedelta(days=1)
           searchresults=ReceptionReport.objects.filter(datecreated__range=(stdate,endate))
           print("C")
      if (not stdate_str and not endate_str):
           searchresults=ReceptionReport.objects.all()
           print("D")
      if bankid:
            searchresults=searchresults.filter(bankid=bankid)
      if engstatus:
            if repstatus == "Pending":
                  searchresults=searchresults.filter(Q(engineer__isnull=True ) | Q(engineer="InProgress"))
            else:
                  searchresults=searchresults.filter(engineer__icontains=engstatus)
      if repstatus:
            if repstatus == "Pending":
                  searchresults=searchresults.filter(Q(reporter__isnull=True ) | Q(reporter="InProgress"))
            else:
                  searchresults=searchresults.filter(reporter__icontains=repstatus)
      if vertical:
            searchresults=searchresults.filter(bankvertical__icontains=vertical)
      # other_field1__icontains  icontains is case insensitive and contains is case sensitive  need from django.db.models import Q 
      
      serializer_result=ReceptionSerializer(searchresults,many=True,context={'request':request}) 
      # print(serializer_result)
      return JsonResponse({'success':True, 'data':serializer_result.data}, status=status.HTTP_200_OK)  

def engineereditedview(request,recid):  
      
      # print(recid) 
      reports = EngineerReport.objects.filter(pk=recid).values(  
      'id', 'applicationnumber', 'name', 'visitinpresence',   
      'bankname', 'casetype', 'add1', 'add2', 'city', 'region',  
      'zip', 'country', 'east', 'west', 'north', 'south',  
      'propertyage', 'landrate', 'Occupant', 'rented',   
      'landmark', 'roadwidth', 'hightensionline',   
      'railwayline', 'nala', 'river', 'pahad',   
      'roadcomesunderroadbinding', 'propertyaccessissue',   
      'othercheck', 'others', 'remark', 'lat', 'lng'  
      ).first()  # Retrieve the single report if you know recid is unique  

# To include floordetails in the result, you can also serialize the data  
      if reports:  
                  floor_details = list(Floordetails.objects.filter(engreportid=recid).values(  
        'floorname', 'floordetail', 'floorarea'))  
                  # dynamicfields = EngDynamicdValue.objects.filter(engreportid=recid)
                  dynamicfields = list(EngDynamicdValue.objects.filter(engreportid=recid).select_related('input_field').order_by('input_field_id').values('input_field_id','value','input_field__label'))
      # Combine the results  
      reports['floors'] = floor_details 
      # dyndata=[]
      # for dynamicfield in dynamicfields: 
            # reports[dynamicfield.input_field.label] = dynamicfield.value
            # dyndata[dynamicfield.input_field.label] = dynamicfield.value
      # keys_list = list(dynamicfields.items())
      reports['dynamicfields'] = dynamicfields
      # keys_list = list(reports.items()) 
      # else:  
      #       engineer_report = None  # Handle case where no record is found  
      # print(reports)
     
#     reports = list(EngineerReport.objects.values('id','applicationnumber', 'name','visitinpresence','bankname','casetype','add1','add2','city',
#                                             'region','zip','country','east','west','north','south',
#                                             'propertyage','landrate','Occupant','rented','landmark','roadwidth','hightensionline',
#                                             'railwayline','nala','river','pahad','roadcomesunderroadbinding','propertyaccessissue','othercheck',
#                                             'others','remark','lat','lng').filter(pk=recid))  
      histories = HistoryEngineerReport.objects.values('id','applicationnumber', 'name','visitinpresence','bankname','casetype','add1','add2','city',
                                                'region','zip','country','east','west','north','south',
                                                'propertyage','landrate','Occupant','rented','landmark','roadwidth','hightensionline',
                                                'railwayline','nala','river','pahad','roadcomesunderroadbinding','propertyaccessissue','othercheck',
                                                'others','remark','lat','lng').filter(engreport_id=recid).order_by('id') 
      # print(keys_list)
      floorhistories=[]
      for history in histories:
            history['floors'] = list(HistoryFloordetails.objects.filter(historyengreportid_id=history.get('id')).values(  
            'floorname', 'floordetail', 'floorarea'))
            # dynamicfields = HistoryEngDynamicdValue.objects.filter(hsengreportid_id=history.get('id'))
            # for dynamicfield in dynamicfields: 
            #       history[dynamicfield.input_field.label] = dynamicfield.value
            history['dynamicfields'] = list(HistoryEngDynamicdValue.objects.filter(hsengreportid_id=history.get('id')).select_related('input_field').order_by('input_field_id').values('input_field_id','value','input_field__label'))
            # print(history)
            floorhistories.append(history)
            # floorhistories += history
      #       floorhistories[history.get('id')] = history
      floorhistories.append(reports) 
      # print(floorhistories)
      #     enghistory = reports.union(histories) 
      
      enghistory =  floorhistories 
      #     serializer_result=EngineerCreateSerializer(enghistory,many=True,context={'request':request})   
      #     print(serializer_result.data)
      return JsonResponse({'success':True, 'data':enghistory}, status=status.HTTP_200_OK)   

# @api_view(['GET'])  
# def application_list(request):  
#     applications = Application.objects.all()  
#     serializer = ApplicationSerializer(applications, many=True)  
#     return Response({"success": True, "data": serializer.data})  